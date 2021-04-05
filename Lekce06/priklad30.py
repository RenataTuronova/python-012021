# program30 - Export do Excelu

# pandas umí uložit data i do Excel sešitu, se kterým se bude uživatelům lépe pracovat. K ukládání do Excelu využívá pandas modul openpyxl, který ale není nainstalován automaticky.
# Nainstaluj si modul openpyxl standardním způsobem, který jsme si ukazovali v lekci.
# Ulož tabulku s cekovými počty vykázaných hodin, kterou jsi vytvořila v příkladu 27 jako Excel soubor. Pokud jsi tento příklad nezpracovala, ulož libovolnou jinou tabulku jako Excel sešit. Výsledný sešit si otevři a zkontroluj. K uložení použij funkci to_excel, se kterou pracuj stejně, jako jsme na lekci pracovali s funkci to_csv.

import pandas
reports = pandas.read_csv("vykazy.csv")
# 1. první pokus uložení do excelu
reports.to_excel("vykazy.xlsx", index=False)

# 2. druhý pokus uložení do excelu (jen na zkoušku :))
employeesPraha = pandas.read_csv("zam_praha.csv")
employeesPlzen = pandas.read_csv("zam_plzeň.csv")
employeesLiberec = pandas.read_csv("zam_liberec.csv")

employeesPraha["mesto"] = "Praha"
employeesPlzen["mesto"] = "Plzeň"
employeesLiberec["mesto"] = "Liberec"

employees = pandas.concat([employeesPlzen, employeesPraha, employeesLiberec], ignore_index=True)
reports = reports.rename(columns={"employee_id": "cislo_zamestnance"})
employeesAndReports = pandas.merge(employees, reports, how="left")

employeesAndReports.to_excel("zamestnanciAVykazy.xlsx", index=False)

# Zkus data z Excelu naopak načíst pomocí funkci read_excel() a ověř, že se soubor načetl v pořádku.
# 1. první pokus načtení dat z Excelu
reportsExcel = pandas.read_excel("vykazy.xlsx")
print(reportsExcel)

# 2. druhý pokus načtení dat v Excelu (jen na zkoušku :))
employeesAndReportsExcel = pandas.read_excel("zamestnanciAVykazy.xlsx")
print(employeesAndReportsExcelExcel)


# ROZŠÍŘENÉ ZADÁNÍ
# Modul openpyxl - zkus pomocí modulu zapsat rozvrh hodin jako tabulku v Excel sešitu. Níže máš program, který obsahuje rozvrh hodin jako dvourozměrný seznam. Vnitřní seznamy obsahují předměty v rozvrhu, jeden vnitřní seznam vždy obsahuje předměty pro daný den.
# Podívej se nejprve na ukázku. Jednoduchý program níže zapíše hodnotu Test do buňky B1 a výsledek uloží souboru rozvrh_hodin.xlsx. Pokud neznáš terminologii Excelu, pak Workbook označuje sešit, tj. celý "soubor". ws se odkazuje na Work Sheet, což je list, tj. jedna "záložka". Náš soubor bude mít jen jeden list, čímž je situace jednoduchá.

# from openpyxl import Workbook

# wb = Workbook()
# ws1 = wb.active
# ws1.title = "rozvrh"
# cislo_radku = 1
# cislo_sloupce = 2

# bunka = ws1.cell(cislo_radku, cislo_sloupce)
# bunka.value = "Test"

# wb.save(filename="rozvrh_hodin.xlsx")

# DOBROVOLNÝ DOPLNĚK ROZŠÍŘENÉHO ZADÁNÍ
# Modul openpyxl je silný v možnostech formátování, které nabízí. Podívej se na rozšířený příklad níže. Ten kromě zápisu hodnoty nastaví buňce šedivou barvu jako pozadí.
# Přehled barev, které můžeš použít, najdeš v dokumentaci.
#
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill
#
# wb = Workbook()
# ws1 = wb.active
# ws1.title = "rozvrh"
# cislo_radku = 1
# cislo_sloupce = 2
#
# bunka = ws1.cell(cislo_radku, cislo_sloupce)
# bunka.value = "Test"
#
# # Určím si barvu
# sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
# # Přiřadím barvu buňce
# bunka.fill = sediva_barva
#
# wb.save(filename="rozvrh_hodin.xlsx")
# Vrať se ke svému programu a zkus vytvořit rozvrh obarvený. Například můžeš zkusit podbarvit oběd šedě a poté třeba nastavit nějakou barvu výchovám, jinou barvu jazykům atd.

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "timetable"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
row = 1
for day in rozvrh_hodin:
  column = 1
  for subject in day:
    cell = ws1.cell(row, column)
    cell.value = subject
    grey = PatternFill("solid", fgColor="00C0C0C0")
    blue = PatternFill("solid", fgColor="00CCFFFF")
    orange = PatternFill("solid", fgColor="00FFCC99")
    green = PatternFill("solid", fgColor="00CCFFCC")
    if subject == "Oběd":
      cell.fill = grey
    elif subject == "Anglický jazyk" or subject == "Český jazyk":
      cell.fill = blue
    elif subject == "Tělesná výchova" or subject == "Hudební výchova" or subject == "Občanská výchova" or subject == "Výtvarná výchova":
      cell.fill = orange
    else:
      cell.fill = green
    column += 1
  row += 1

wb.save(filename="rozvrh_hodin.xlsx")

